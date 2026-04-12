import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
from openpyxl.utils import get_column_letter

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Lịch trình Đắk Lắk"

# ── Colors ──
DARK_BG = "1C1917"
WARM_BG = "292524"
RED_ACCENT = "DC2626"
GOLD = "F59E0B"
GREEN = "22C55E"
WHITE = "FAFAF9"
MUTED = "A8A29E"
SUBTLE = "78716C"

# ── Fonts ──
font_title = Font(name="Arial", size=18, bold=True, color=WHITE)
font_subtitle = Font(name="Arial", size=11, color=MUTED, italic=True)
font_day = Font(name="Arial", size=14, bold=True, color=WHITE)
font_header = Font(name="Arial", size=10, bold=True, color=WHITE)
font_time = Font(name="Courier New", size=11, bold=True, color=GOLD)
font_place = Font(name="Arial", size=11, bold=True, color=WHITE)
font_place_star = Font(name="Arial", size=11, bold=True, color=GOLD)
font_desc = Font(name="Arial", size=10, color=MUTED)
font_tag = Font(name="Arial", size=9, color=SUBTLE)
font_duration = Font(name="Courier New", size=10, bold=True, color="60A5FA")
font_cost = Font(name="Arial", size=10, bold=True, color=GREEN)
font_cost_total = Font(name="Arial", size=11, bold=True, color=GOLD)
font_dist = Font(name="Arial", size=9, color=MUTED, italic=True)
font_check = Font(name="Arial", size=9, color=MUTED)
font_section = Font(name="Arial", size=9, bold=True, color=RED_ACCENT)
font_note = Font(name="Arial", size=10, color=GOLD, italic=True)

# ── Fills ──
fill_dark = PatternFill(start_color=DARK_BG, end_color=DARK_BG, fill_type="solid")
fill_warm = PatternFill(start_color=WARM_BG, end_color=WARM_BG, fill_type="solid")
fill_gold = PatternFill(start_color="3D2800", end_color="3D2800", fill_type="solid")
fill_header = PatternFill(start_color="3F3A36", end_color="3F3A36", fill_type="solid")
fill_total = PatternFill(start_color="1A2E1A", end_color="1A2E1A", fill_type="solid")

# ── Border ──
thin_border = Border(
    left=Side(style='thin', color=SUBTLE),
    right=Side(style='thin', color=SUBTLE),
    top=Side(style='thin', color=SUBTLE),
    bottom=Side(style='thin', color=SUBTLE)
)

# ── Alignment ──
align_center = Alignment(horizontal='center', vertical='center', wrap_text=True)
align_left = Alignment(horizontal='left', vertical='center', wrap_text=True)
align_top = Alignment(horizontal='left', vertical='top', wrap_text=True)
align_right = Alignment(horizontal='right', vertical='center', wrap_text=True)

# ── Column widths ──
# A=STT, B=Giờ, C=Thời lượng, D=Địa điểm, E=Mô tả, F=Chi phí, G=Khoảng cách, H=Phân loại, I=Checklist
col_widths = {1: 4, 2: 9, 3: 12, 4: 30, 5: 50, 6: 18, 7: 22, 8: 16, 9: 34}
for col, width in col_widths.items():
    ws.column_dimensions[get_column_letter(col)].width = width

# ── Fill background ──
for row in range(1, 65):
    for col in range(1, 10):
        ws.cell(row=row, column=col).fill = fill_dark

# ══════════════════════════════════════
# TITLE
# ══════════════════════════════════════
row = 2
ws.merge_cells('B2:I2')
c = ws.cell(row=2, column=2, value="LỊCH TRÌNH ĐẮK LẮK — NGHIÊN CỨU DỆT THỔ CẨM Ê ĐÊ")
c.font = font_title
c.alignment = align_center

row = 3
ws.merge_cells('B3:I3')
c = ws.cell(row=3, column=2, value="2 ngày · 2 người · Xe máy thuê · Khách sạn Song Anh (Hai Bà Trưng) · Thứ 6–7")
c.font = font_subtitle
c.alignment = align_center

# ══════════════════════════════════════
# HEADER
# ══════════════════════════════════════
row = 5
headers = ["", "Giờ", "Thời lượng", "Địa điểm", "Mô tả & Ghi chú", "Chi phí (2 người)", "Khoảng cách", "Phân loại", "Checklist nghiên cứu"]
for col_idx, header in enumerate(headers, 1):
    cell = ws.cell(row=row, column=col_idx, value=header)
    cell.font = font_header
    cell.fill = fill_header
    cell.alignment = align_center
    cell.border = thin_border

# ══════════════════════════════════════
# DATA — updated itinerary
# ══════════════════════════════════════
# Format: (type, giờ, thời_lượng, địa_điểm, mô_tả, chi_phí_str, khoảng_cách, phân_loại, checklist, is_star)

itinerary = [
    # ─── NGÀY 1 ───
    ("day", "NGÀY 1 — THỨ SÁU", "", "Bya Hoa · Bảo tàng · Ako Dhong · Cà Phê", "", "", "", "", "", False),

    ("row", "07:00", "1.5h", "✈️ Đến Buôn Ma Thuột",
     "Đến sân bay BMT. Nhận xe máy thuê (~150k/ngày).\nCheck-in Song Anh Hotel, đường Hai Bà Trưng.",
     "Xe máy: 300k (2 ngày)\nHotel: 350k/đêm × 1 = 350k",
     "Sân bay → Hotel\n~8 km, 15 phút",
     "Di chuyển", "", False),

    ("row", "08:30", "2h", "⭐ Bya Hoa — Thổ cẩm đương đại",
     "Thương hiệu thổ cẩm Ê Đê. Phỏng vấn chủ thương hiệu: truyền thống + hiện đại.\n🎯 Bảo tồn nghề dệt qua mô hình kinh doanh.",
     "Miễn phí (tham quan)",
     "Hotel → Bya Hoa\n~3 km, 8 phút",
     "⭐ Phỏng vấn",
     "☐ Phỏng vấn founder\n☐ Chụp sản phẩm & quy trình\n☐ Hỏi nguồn nguyên liệu\n☐ Tìm hiểu đầu ra", True),

    ("row", "10:30", "1.5h", "⭐ Bảo tàng tỉnh Đắk Lắk",
     "Khu trưng bày thổ cẩm Ê Đê, M'nông. Khung dệt, trang phục cổ, thông tin học thuật.\n🎯 Hiểu rõ nguồn gốc & cấu trúc hoa văn.",
     "Vé: ~20k × 2 = 40k",
     "Bya Hoa → BT Đắk Lắk\n~2 km, 5 phút",
     "⭐ Nghiên cứu",
     "☐ Ghi chép cấu trúc khung dệt\n☐ Chụp chi tiết hoa văn\n☐ Note bảng thuyết minh\n☐ So sánh Ê Đê vs M'nông", True),

    ("row", "12:00", "1.5h", "🍜 Ăn trưa",
     "Đặc sản Tây Nguyên — cơm lam, gà nướng, canh lá. Quán gần bảo tàng.",
     "~80k/người × 2 = 160k",
     "BT → quán ăn\n~1 km",
     "Ẩm thực", "", False),

    ("row", "13:30", "2h", "📍 Buôn Ako Dhong",
     "Làng cổ Ê Đê nổi tiếng trong lòng TP. Nhà dài truyền thống, văn hóa Ê Đê. Dệt thường diễn ra cuối tuần.",
     "Miễn phí",
     "Quán ăn → Ako Dhong\n~3 km, 8 phút",
     "Văn hóa",
     "☐ Chụp nhà dài & kiến trúc\n☐ Hỏi vai trò dệt trong đời sống\n☐ Tìm nghệ nhân tại buôn", False),

    ("row", "15:30", "2h", "☕ BT Thế Giới Cà Phê",
     "Thư giãn. Cà phê đặc sản BMT. Chỉnh lý ghi chép buổi sáng.",
     "Vé: ~80k × 2 = 160k\nCà phê: ~40k × 2 = 80k",
     "Ako Dhong → BT Cà Phê\n~4 km, 10 phút",
     "Nghỉ ngơi", "", False),

    ("row", "18:00", "1h", "🍜 Ăn tối",
     "Ăn tối đặc sản. Bún đỏ, cơm niêu, lẩu lá rừng.",
     "~100k/người × 2 = 200k",
     "BT Cà Phê → Quán\n~3 km",
     "Ẩm thực", "", False),

    ("row", "20:00", "—", "🌙 Về hotel — Tổng hợp",
     "Tổng hợp ghi chú, sắp xếp ảnh. Chuẩn bị câu hỏi phỏng vấn ngày 2 tại HTX Tơng Jú.",
     "—",
     "Quán → Hotel\n~2 km",
     "Chuẩn bị", "", False),

    # ─── NGÀY 2 ───
    ("day", "NGÀY 2 — THỨ BẢY", "", "HTX Tơng Jú · Về Nha Trang", "", "", "", "", "", False),

    ("row", "07:00", "1h", "☕ Sáng — Cà phê & Chuẩn bị",
     "Ăn sáng + cà phê. Kiểm tra thiết bị, pin, bộ nhớ. Xác nhận lịch hẹn HTX Tơng Jú.",
     "Ăn sáng: ~40k × 2 = 80k\nCà phê: ~30k × 2 = 60k",
     "—",
     "Chuẩn bị", "", False),

    ("row", "08:00", "4h", "⭐⭐ HTX Dệt Thổ Cẩm Tơng Jú",
     "ĐIỂM QUAN TRỌNG NHẤT — Dệt là nghề chính, HTX có tổ chức. Quan sát toàn bộ quy trình, phỏng vấn nghệ nhân. Dành ~4 tiếng.\n🎯 Dành đủ thời gian!",
     "Miễn phí (tham quan)",
     "Hotel → Tơng Jú (Ea Kao)\n~13 km, 25 phút",
     "⭐⭐ NC chính",
     "☐ Quan sát quy trình dệt đầu→cuối\n☐ PV nghệ nhân (kỹ thuật, hoa văn)\n☐ PV ban quản lý HTX\n☐ Quay video quy trình\n☐ Chụp khung dệt & nguyên liệu\n☐ Hỏi ý nghĩa hoa văn\n☐ Đầu ra sản phẩm HTX", True),

    ("row", "12:00", "1.5h", "🍜 Ăn trưa",
     "Ăn gần Tơng Jú hoặc quay về trung tâm.",
     "~80k × 2 = 160k",
     "Tơng Jú → Quán\n~5–13 km",
     "Ẩm thực", "", False),

    ("row", "13:30", "2h", "🔄 Dự phòng / Bổ sung",
     "Quay lại Bya Hoa hoặc Ako Dhong nếu cần bổ sung. Hoặc mua sắm đặc sản Tây Nguyên (cà phê, mật ong, hạt điều).",
     "Mua sắm: ~200k (tùy)",
     "Tùy điểm\n~3–5 km",
     "Tự do", "", False),

    ("row", "15:30", "1h", "📦 Thu dọn & Trả xe",
     "Về hotel trả phòng. Trả xe máy. Sắp xếp hành lý.",
     "Xăng xe 2 ngày: ~80k",
     "→ Hotel → Trả xe\n~3 km",
     "Logistics", "", False),

    ("row", "17:30", "1h", "🍜 Ăn tối nhẹ trước khi đi",
     "Ăn nhẹ gần bến xe trước khi lên xe.",
     "~60k × 2 = 120k",
     "Hotel → Quán\n~2 km",
     "Ẩm thực", "", False),

    ("row", "18:30", "5–6h", "🚌 Xe buýt → Nha Trang",
     "Xe giường nằm BMT → Nha Trang (~5–6 tiếng). Tổng hợp ghi chép trên xe.",
     "Vé xe: ~250k × 2 = 500k",
     "Quán → Bến xe\n~3 km, 8 phút",
     "Di chuyển", "", False),
]

# ══════════════════════════════════════
# WRITE DATA
# ══════════════════════════════════════
row = 6
for item in itinerary:
    item_type = item[0]

    if item_type == "day":
        # Day header
        ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=9)
        cell = ws.cell(row=row, column=2, value=f"{item[1]}   ·   {item[3]}")
        cell.font = font_day
        cell.alignment = align_center
        for col in range(1, 10):
            ws.cell(row=row, column=col).fill = fill_warm
        ws.row_dimensions[row].height = 35
        row += 1
        continue

    is_star = item[9]
    row_fill = fill_gold if is_star else fill_dark

    # Col A — empty
    ws.cell(row=row, column=1).fill = row_fill

    # Col B — Giờ
    cell = ws.cell(row=row, column=2, value=item[1])
    cell.font = font_time
    cell.alignment = align_center
    cell.fill = row_fill
    cell.border = thin_border

    # Col C — Thời lượng
    cell = ws.cell(row=row, column=3, value=item[2])
    cell.font = font_duration
    cell.alignment = align_center
    cell.fill = row_fill
    cell.border = thin_border

    # Col D — Địa điểm
    cell = ws.cell(row=row, column=4, value=item[3])
    cell.font = font_place_star if is_star else font_place
    cell.alignment = align_left
    cell.fill = row_fill
    cell.border = thin_border

    # Col E — Mô tả
    cell = ws.cell(row=row, column=5, value=item[4])
    cell.font = font_desc
    cell.alignment = align_top
    cell.fill = row_fill
    cell.border = thin_border

    # Col F — Chi phí
    cell = ws.cell(row=row, column=6, value=item[5])
    cell.font = font_cost
    cell.alignment = align_top
    cell.fill = row_fill
    cell.border = thin_border

    # Col G — Khoảng cách
    cell = ws.cell(row=row, column=7, value=item[6])
    cell.font = font_dist
    cell.alignment = align_top
    cell.fill = row_fill
    cell.border = thin_border

    # Col H — Phân loại
    cell = ws.cell(row=row, column=8, value=item[7])
    cell.font = font_section if "⭐" in item[7] else font_tag
    cell.alignment = align_center
    cell.fill = row_fill
    cell.border = thin_border

    # Col I — Checklist
    cell = ws.cell(row=row, column=9, value=item[8])
    cell.font = font_check
    cell.alignment = align_top
    cell.fill = row_fill
    cell.border = thin_border

    # Row height
    lines = max(
        item[4].count('\n') + 1,
        item[5].count('\n') + 1,
        item[6].count('\n') + 1,
        item[8].count('\n') + 1,
        1
    )
    ws.row_dimensions[row].height = max(32, lines * 17)
    row += 1

# ══════════════════════════════════════
# TỔNG CHI PHÍ
# ══════════════════════════════════════
row += 1
ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=9)
cell = ws.cell(row=row, column=2, value="TỔNG CHI PHÍ ƯỚC TÍNH (2 NGƯỜI)")
cell.font = font_day
cell.alignment = align_center
for col in range(1, 10):
    ws.cell(row=row, column=col).fill = fill_warm
ws.row_dimensions[row].height = 35
row += 1

costs = [
    ("🏨 Khách sạn", "Song Anh Hotel (Hai Bà Trưng) × 1 đêm", "350,000đ"),
    ("🏍️ Thuê xe máy", "~150k/ngày × 2 ngày", "300,000đ"),
    ("⛽ Xăng xe", "~40km/ngày, 2 ngày", "80,000đ"),
    ("🍜 Ăn uống (6 bữa)", "2 sáng + 2 trưa + 2 tối, ×2 người", "780,000đ"),
    ("☕ Cà phê", "2 lần, ×2 người", "140,000đ"),
    ("🎟️ Vé tham quan", "BT Đắk Lắk + BT Cà Phê, ×2 người", "200,000đ"),
    ("🚌 Xe BMT → Nha Trang", "Xe giường nằm ×2 người", "500,000đ"),
    ("🛍️ Mua sắm đặc sản", "Tùy chọn", "~200,000đ"),
    ("", "", ""),
    ("💰 TỔNG CỘNG", "", "~2,550,000đ"),
]

for c_item in costs:
    if c_item[0] == "":
        row += 1
        continue

    is_total = "TỔNG" in c_item[0]
    rf = fill_total if is_total else fill_dark

    cell = ws.cell(row=row, column=1)
    cell.fill = rf

    ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=4)
    cell = ws.cell(row=row, column=2, value=c_item[0])
    cell.font = font_cost_total if is_total else font_place
    cell.alignment = align_left
    cell.fill = rf
    cell.border = thin_border
    for c in [3, 4]:
        ws.cell(row=row, column=c).fill = rf
        ws.cell(row=row, column=c).border = thin_border

    ws.merge_cells(start_row=row, start_column=5, end_row=row, end_column=7)
    cell = ws.cell(row=row, column=5, value=c_item[1])
    cell.font = font_desc
    cell.alignment = align_left
    cell.fill = rf
    cell.border = thin_border
    for c in [6, 7]:
        ws.cell(row=row, column=c).fill = rf
        ws.cell(row=row, column=c).border = thin_border

    ws.merge_cells(start_row=row, start_column=8, end_row=row, end_column=9)
    cell = ws.cell(row=row, column=8, value=c_item[2])
    cell.font = font_cost_total if is_total else font_cost
    cell.alignment = align_right
    cell.fill = rf
    cell.border = thin_border
    ws.cell(row=row, column=9).fill = rf
    ws.cell(row=row, column=9).border = thin_border

    ws.row_dimensions[row].height = 30
    row += 1

# ══════════════════════════════════════
# GHI CHÚ
# ══════════════════════════════════════
row += 1
ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=9)
cell = ws.cell(row=row, column=2, value="📌 GHI CHÚ: Chi phí ước tính tham khảo, có thể thay đổi. Khoảng cách tính từ Google Maps. Đi xe máy nên cộng thêm 5-10 phút mỗi chặng.")
cell.font = font_note
cell.alignment = align_left
for col in range(1, 10):
    ws.cell(row=row, column=col).fill = fill_dark

# ── Save ──
output_path = "/Users/lequan275/Downloads/DAKLAK TIMELINE/Lich_trinh_DakLak.xlsx"
wb.save(output_path)
print(f"✅ Đã tạo: {output_path}")
